"""
Content detection + extraction pipeline.
Design goal: detect real content type first, then choose parser.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Optional
import io
import json
import mimetypes
import re
import zipfile
import xml.etree.ElementTree as ET

from loguru import logger

try:
    import magic  # type: ignore
    _MAGIC_OK = True
except Exception:
    magic = None
    _MAGIC_OK = False


@dataclass
class DetectedType:
    mime_type: str
    group: str
    source: str
    confidence: float = 0.0
    extension_hint: str = ""
    parser_hint: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ExtractionResult:
    detected_type: str
    parser: str
    text: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)
    confidence: float = 0.0
    needs_ocr: bool = False
    encrypted: bool = False
    truncated: bool = False
    error: str = ""


class ContentProcessor:
    """Detect real type then extract with parser registry."""

    def __init__(self, max_text_length: int = 10_000):
        self.max_text_length = int(max_text_length)
        self._magic = None
        if _MAGIC_OK:
            try:
                self._magic = magic.Magic(mime=True)  # type: ignore[attr-defined]
            except Exception as e:
                logger.debug(f"magic init failed: {e}")
                self._magic = None

    def detect_file_type(self, file_path: Path) -> DetectedType:
        ext = file_path.suffix.lower()
        header = self._read_header(file_path, 8192)
        magic_mime = self._detect_magic_mime(file_path)
        container_mime = self._detect_container_type(file_path, header, ext)

        if container_mime:
            mime_t = container_mime
            source = "container-aware"
            conf = 0.95
        elif magic_mime:
            mime_t = magic_mime
            source = "magic-bytes"
            conf = 0.9
        else:
            mime_t = self._guess_text_or_mime(file_path, header, ext)
            source = "heuristic"
            conf = 0.65

        group = self._group_from_mime(mime_t)
        return DetectedType(
            mime_type=mime_t,
            group=group,
            source=source,
            confidence=conf,
            extension_hint=ext,
            parser_hint=group,
        )

    def extract_content(self, file_path: Path, detected: DetectedType) -> ExtractionResult:
        group = detected.group
        if group == "text":
            return self._extract_text_like(file_path, detected)
        if group == "docx":
            return self._extract_docx(file_path, detected)
        if group == "xlsx":
            return self._extract_xlsx(file_path, detected)
        if group == "pdf":
            return self._extract_pdf(file_path, detected)
        if group == "image":
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="image-placeholder",
                needs_ocr=True,
                confidence=detected.confidence,
                metadata={"group": group},
            )
        if group == "archive":
            return self._extract_archive_metadata(file_path, detected)
        return self._extract_with_tika_fallback(file_path, detected)

    def _read_header(self, file_path: Path, n: int) -> bytes:
        try:
            with open(file_path, "rb") as f:
                return f.read(n)
        except Exception:
            return b""

    def _detect_magic_mime(self, file_path: Path) -> str:
        if not self._magic:
            return ""
        try:
            return str(self._magic.from_file(str(file_path)) or "").strip()
        except Exception:
            return ""

    def _detect_container_type(self, file_path: Path, header: bytes, ext: str) -> str:
        if header.startswith(b"%PDF"):
            return "application/pdf"
        if header.startswith(b"\x89PNG"):
            return "image/png"
        if header.startswith(b"\xff\xd8\xff"):
            return "image/jpeg"
        if header.startswith((b"GIF87a", b"GIF89a")):
            return "image/gif"
        if header.startswith(b"PK\x03\x04"):
            try:
                with zipfile.ZipFile(file_path, "r") as zf:
                    names = set(zf.namelist())
                    if "[Content_Types].xml" in names:
                        if any(n.startswith("word/") for n in names):
                            return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                        if any(n.startswith("xl/") for n in names):
                            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        if any(n.startswith("ppt/") for n in names):
                            return "application/vnd.openxmlformats-officedocument.presentationml.presentation"
                    return "application/zip"
            except Exception:
                return "application/zip"
        if ext == ".csv":
            return "text/csv"
        return ""

    def _guess_text_or_mime(self, file_path: Path, header: bytes, ext: str) -> str:
        if self._looks_like_text(header):
            if ext == ".csv":
                return "text/csv"
            if ext in {".json"}:
                return "application/json"
            if ext in {".xml"}:
                return "application/xml"
            return "text/plain"
        guessed = mimetypes.guess_type(str(file_path))[0]
        return guessed or "application/octet-stream"

    def _looks_like_text(self, header: bytes) -> bool:
        if not header:
            return False
        sample = header[:2048]
        if b"\x00" in sample:
            return False
        printable = sum((32 <= b <= 126) or b in (9, 10, 13) for b in sample)
        ratio = printable / max(len(sample), 1)
        return ratio >= 0.85

    def _group_from_mime(self, mime_t: str) -> str:
        m = (mime_t or "").lower()
        if m.startswith("text/") or m in {"application/json", "application/xml"}:
            return "text"
        if m == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return "docx"
        if m == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return "xlsx"
        if m == "application/pdf":
            return "pdf"
        if m.startswith("image/"):
            return "image"
        if "zip" in m or "rar" in m or "7z" in m:
            return "archive"
        return "unknown"

    def _extract_text_like(self, file_path: Path, detected: DetectedType) -> ExtractionResult:
        for enc in ("utf-8", "utf-16", "latin-1"):
            try:
                with open(file_path, "r", encoding=enc, errors="ignore") as f:
                    data = f.read(self.max_text_length + 1)
                return ExtractionResult(
                    detected_type=detected.mime_type,
                    parser=f"plain-text:{enc}",
                    text=data[: self.max_text_length],
                    truncated=len(data) > self.max_text_length,
                    confidence=detected.confidence,
                    metadata={"group": detected.group},
                )
            except Exception:
                continue
        return ExtractionResult(
            detected_type=detected.mime_type,
            parser="plain-text",
            confidence=0.0,
            error="cannot-read-text",
            metadata={"group": detected.group},
        )

    def _extract_docx(self, file_path: Path, detected: DetectedType) -> ExtractionResult:
        try:
            from docx import Document  # type: ignore
            doc = Document(str(file_path))
            parts = [p.text for p in doc.paragraphs if p.text]
            text = "\n".join(parts)
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="python-docx",
                text=text[: self.max_text_length],
                truncated=len(text) > self.max_text_length,
                confidence=detected.confidence,
                metadata={"paragraphs": len(parts), "group": detected.group},
            )
        except Exception:
            return self._extract_docx_xml_fallback(file_path, detected)

    def _extract_docx_xml_fallback(self, file_path: Path, detected: DetectedType) -> ExtractionResult:
        try:
            with zipfile.ZipFile(file_path, "r") as zf:
                xml_bytes = zf.read("word/document.xml")
            root = ET.fromstring(xml_bytes)
            text_nodes = [n.text for n in root.iter() if n.tag.endswith("}t") and n.text]
            text = "\n".join(text_nodes)
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="docx-xml-fallback",
                text=text[: self.max_text_length],
                truncated=len(text) > self.max_text_length,
                confidence=max(0.5, detected.confidence - 0.2),
                metadata={"group": detected.group},
            )
        except Exception as e:
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="docx-xml-fallback",
                confidence=0.0,
                error=str(e),
                metadata={"group": detected.group},
            )

    def _extract_xlsx(self, file_path: Path, detected: DetectedType) -> ExtractionResult:
        try:
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            chunks = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    vals = [str(v) for v in row if v not in (None, "")]
                    if vals:
                        chunks.append(",".join(vals))
                        if sum(len(x) for x in chunks) > self.max_text_length:
                            break
                if sum(len(x) for x in chunks) > self.max_text_length:
                    break
            text = "\n".join(chunks)
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="openpyxl",
                text=text[: self.max_text_length],
                truncated=len(text) > self.max_text_length,
                confidence=detected.confidence,
                metadata={"sheets": len(wb.worksheets), "group": detected.group},
            )
        except Exception as e:
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="openpyxl",
                confidence=0.0,
                error=str(e),
                metadata={"group": detected.group},
            )

    def _extract_pdf(self, file_path: Path, detected: DetectedType) -> ExtractionResult:
        try:
            from pypdf import PdfReader  # type: ignore
            reader = PdfReader(str(file_path))
            chunks = []
            for p in reader.pages:
                chunks.append(p.extract_text() or "")
                if sum(len(x) for x in chunks) > self.max_text_length:
                    break
            text = "\n".join(chunks)
            needs_ocr = len(text.strip()) < 40
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="pypdf",
                text=text[: self.max_text_length],
                truncated=len(text) > self.max_text_length,
                confidence=detected.confidence,
                needs_ocr=needs_ocr,
                metadata={"pages": len(reader.pages), "group": detected.group},
            )
        except Exception as e:
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="pypdf",
                confidence=0.0,
                error=str(e),
                needs_ocr=True,
                metadata={"group": detected.group},
            )

    def _extract_archive_metadata(self, file_path: Path, detected: DetectedType) -> ExtractionResult:
        encrypted = False
        entries = 0
        if file_path.suffix.lower() == ".zip":
            try:
                with zipfile.ZipFile(file_path, "r") as zf:
                    infos = zf.infolist()
                    entries = len(infos)
                    encrypted = any(i.flag_bits & 0x1 for i in infos)
            except Exception:
                pass
        return ExtractionResult(
            detected_type=detected.mime_type,
            parser="archive-meta",
            text="",
            confidence=detected.confidence,
            encrypted=encrypted,
            metadata={"entries": entries, "group": detected.group},
        )

    def _extract_with_tika_fallback(self, file_path: Path, detected: DetectedType) -> ExtractionResult:
        try:
            from tika import parser  # type: ignore
            parsed = parser.from_file(str(file_path))
            content = (parsed or {}).get("content") or ""
            meta = (parsed or {}).get("metadata") or {}
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="tika",
                text=str(content)[: self.max_text_length],
                truncated=len(str(content)) > self.max_text_length,
                confidence=max(0.5, detected.confidence),
                metadata={"tika_metadata_keys": list(meta.keys())[:20], "group": detected.group},
            )
        except Exception as e:
            return ExtractionResult(
                detected_type=detected.mime_type,
                parser="unknown",
                confidence=0.0,
                error=f"no-parser:{e}",
                metadata={"group": detected.group},
            )
